from tkinter import filedialog as fd
from tkinter import ttk
import tkinter.messagebox as mb
import tkinter as tk
import pandas as pd
import os
import numpy as np
from datetime import datetime, timedelta
from pandas.io.excel import ExcelWriter
import openpyxl as ox
from openpyxl.styles import (
                        PatternFill, Border, Side, 
                        Alignment, Font, GradientFill
                        )
import os.path


def read_excel():
    db_dowload = fd.askopenfilename(title='Выбор файла с выгрузкой', initialdir=r'C:\Users\User\Емиас\Загрузка вакцин из ЛК\Выгрузки')
    
    progress_label.config(text='Идёт чтение файла...')
    progress_label.update()
    
    df1 = pd.read_excel(db_dowload, sheet_name='Main', dtype=str)
    
    progress_label.config(text='Файл загружен!')
    progress_label.update()

    #Чтение дат
    df1['BIRTH_DATE'] = pd.to_datetime(df1['BIRTH_DATE'])
    df1['DOCUMENT_CREATED'] = pd.to_datetime(df1['DOCUMENT_CREATED']).dt.date

    #сортировка по дате создания докумена
    df1_sort = df1.sort_values(['DOCUMENT_CREATED'], ascending=[True])

    #расчет возраста
    today = pd.Timestamp(datetime.now())
    df1_sort['Age'] = (today - df1_sort['BIRTH_DATE']) / pd.Timedelta(days=365.25)

    #отбор нужных значений
    df1_sort_adult = df1_sort[df1_sort.STATUS == 'НЕ ОБРАБОТАН врачом']
    df1_sort_adult = df1_sort_adult[df1_sort_adult.Age > 18.25]
    df1_sort_adult = df1_sort_adult[df1_sort_adult.FAMILY !='Тест']
    df1_sort_adult = df1_sort_adult[df1_sort_adult.FAMILY !='Тестовый']
    

    #считаем количество загруженных документов
    df_unic_doc = df1_sort_adult.drop_duplicates (subset=['DOCUMENT_SUBMISSION_SET_ID'])
    df_oms = df_unic_doc.POL_NUMBER.value_counts()

    #отбираем только уникальные ОМС
    df1_sort_adult_unic = df1_sort_adult.drop_duplicates (subset=['POL_NUMBER'])

    #создаем копию
    df1_sort_adult_unic_count = df1_sort_adult_unic.copy()

    #добавляем количество загруженных документов
    df1_sort_adult_unic_count['Количество необработанных документов'] = df1_sort_adult_unic['POL_NUMBER'].map(df_oms)

    #Сохраняем оригинал выгрузки
    df1_sort_adult_unic_count.to_excel(os.path.dirname(db_dowload) + '\Выборка ' + os.path.basename(db_dowload))
        
    # Подстановка коротких названий МО в выгрузку
    df_xlsm = df1_sort_adult_unic_count.loc[:,('POL_NUMBER', 'LPU_NAME', 'MAIN_LPU_NAME', 'DOCUMENT_CREATED', 'Количество необработанных документов')]
    df_base = pd.read_excel(file_xlsm, sheet_name='База', dtype=str)
    
    df_unic_mo = df_base.drop_duplicates (subset=['МО прикрепления']).dropna(subset=['МО прикрепления'])
    df_mo = pd.Series(df_unic_mo['Сокращ. МО'].to_list(), df_unic_mo['МО прикрепления'])
    df_unic_mo_oms = df_base.drop_duplicates (subset=['ОМС']).dropna(subset=['ОМС'])
    df_mo = df_mo.append(pd.Series(df_unic_mo_oms['Сокращ. МО'].to_list(), df_unic_mo_oms['ОМС']))

    df_unic_fil = df_base.drop_duplicates (subset=['Филиал прикрепления']).dropna(subset=['Филиал прикрепления'])
    df_fil = pd.Series(df_unic_fil['Сокращ. Филиал'].to_list(), df_unic_fil['Филиал прикрепления'])
    df_unic_fil_oms = df_base.drop_duplicates (subset=['ОМС']).dropna(subset=['ОМС'])
    df_fil = df_fil.append(pd.Series(df_unic_fil_oms['Сокращ. Филиал'].to_list(), df_unic_mo_oms['ОМС']))

    df_xlsm_count = df_xlsm.copy()
    df_xlsm_count['МО прикрепления'] = df_xlsm['MAIN_LPU_NAME']
    df_xlsm_count['МО прикрепления'].fillna(df_xlsm_count['POL_NUMBER'], inplace=True)
    df_xlsm_count['МО прикрепления'] = df_xlsm_count['МО прикрепления'].map(df_mo)
    hand_input = False
    if len(df_xlsm_count['МО прикрепления'].isna()) != 0:
        hand_input = True

    df_xlsm_count['Филиал прикрепления'] = df_xlsm['LPU_NAME']
    df_xlsm_count['Филиал прикрепления'].fillna(df_xlsm_count['POL_NUMBER'], inplace=True)
    df_xlsm_count['Филиал прикрепления'] = df_xlsm_count['Филиал прикрепления'].map(df_fil)
    if len(df_xlsm_count['МО прикрепления'].isna()) != 0:
        hand_input = True

    df_xlsm_count = df_xlsm_count[['POL_NUMBER', 'Филиал прикрепления', 'МО прикрепления', 'DOCUMENT_CREATED', 'Количество необработанных документов', 'LPU_NAME', 'MAIN_LPU_NAME']]
    df_xlsm_count['Количество необработанных документов'] = df_xlsm_count['Количество необработанных документов'].astype (int)


    #Заполнение базы выгрузкой
    wb = ox.load_workbook(filename=file_xlsm, read_only=False, keep_vba=True)
    name_list = str(datetime.now().date().strftime("%d.%m.%y"))
    xls = pd.ExcelFile(file_xlsm)
    for lists_xls in xls.sheet_names:
            if '.' in lists_xls:
                wb[lists_xls].title = name_list
            
    wb[name_list].delete_rows(2, 10000)
    
    for ir in range(0, len(df_xlsm_count)):
        for ic in range(0, len(df_xlsm_count.iloc[ir])):
            wb[name_list].cell(2 + ir, 2 + ic).value = df_xlsm_count.iloc[ir][ic]

    wb.save(file_xlsm)
    wb.close()

    if hand_input:
        mb.showerror('Требуется ручное вмешательство', 'Добавьте новые ОМС в базу руками')
        os.startfile(file_xlsm)
    else:
        sep_base_mo()


def sep_base_mo():
    #Проверка и создание папки
    today = pd.Timestamp(datetime.now()).date().strftime("%d.%m.%y")
    folder = os.path.dirname(file_xlsm) + '\\Разделенные выгрузки\\' + today + '\\'
    if not os.path.exists(folder):
        if not os.path.exists(os.path.dirname(file_xlsm) + '\\Разделенные выгрузки\\'):
            os.mkdir(os.path.dirname(file_xlsm) + '\\Разделенные выгрузки\\')
        else:
            if not os.path.exists(os.path.dirname(file_xlsm) + '\\Разделенные выгрузки\\' + today):
                os.mkdir(os.path.dirname(file_xlsm) + '\\Разделенные выгрузки\\' + today)
        
    # чтение исходного файла
    df_base_mo = pd.read_excel(file_xlsm, sheet_name=today, dtype=str)
    
    df_base_mo['Количество необработанных документов'] = df_base_mo['Количество необработанных документов'].astype (int)
    # разделение датафрейма по столбцу "МО"
    grouped = df_base_mo.iloc[:, 1:6].groupby('МО')

    # сохранение каждой группы как отдельный файл
    for name, group in grouped:
        group.to_excel(f'{folder}{name} {today}.xlsx', index=False)

    # Редактирование созданных файлов
    name_mo = df_base_mo['МО'].drop_duplicates()
    for name in name_mo:
        file_MO = f'{folder}{name} {today}.xlsx'
        wb = ox.load_workbook(filename=file_MO, read_only=False)
        ws = wb.active
        for ii in range(5):
            ws.cell(1, 1 + ii).value = ['ОМС', 'Филиал', 'МО', 'Дата загрузки \nдокумента', 'Количество \nнеобработанных \nдокументов'][ii]

        for i in ('ABCDE'):
            ws[i + '1'].font = Font(bold=True)
            ws[i + '1'].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            ws[i + '1'].fill = PatternFill('solid', fgColor='FFFF00')
        ws.column_dimensions['A'].width = 17
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
    
        wb.save(file_MO)
        wb.close()


    #создание отчета
    wb = ox.load_workbook(filename=file_xlsm, read_only=False, keep_vba=True)
    ws = wb['Отчет']
    last_row = ws.max_row
    ws.insert_rows(last_row)
    thins = Side(border_style="thin", color="000000")
    for col in range(1, 8):
        if col == 6 or col == 7:
            ws.cell(row=last_row, column=col).font = Font(bold=True) 
        ws.cell(row=last_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=last_row, column=col).border = Border(top=thins, bottom=thins, left=thins, right=thins)

    ws.cell(last_row, 1).value = datetime.now().date()
    ws.cell(last_row, 1).number_format = 'DD.MM.YYYY'
    ws.cell(last_row, 4).value = len(df_base_mo)
    ws.cell(last_row, 5).value = len(name_mo)
    ws.cell(last_row, 6).value = f'=D{last_row}+B{last_row}'
    ws.cell(last_row, 7).value = f'=E{last_row}+C{last_row}'
    
    ws.cell(last_row + 1, 2).value = f'=SUM(B5:B{last_row})'
    ws.cell(last_row + 1, 4).value = f'=SUM(D5:D{last_row})'
    ws.cell(last_row + 1, 6).value = f'=SUM(F5:F{last_row})'

    wb.save(file_xlsm)
    wb.close()


    os.startfile(os.path.dirname(file_MO))
    os.startfile(file_xlsm)
    print("Работа сделана!!")

    root.destroy()


#Проверка наличие базы
file_xlsm =r'C:\Users\User\Емиас\Загрузка вакцин из ЛК\Выгрузка ЛК база с рассылкой.xlsm'
if os.path.isfile(file_xlsm):
    print ('Файл с базой найден')
else:
    print ('Файл с базой НЕ найден!!!')
    print ('Выберите файл с базой для рассыллки ')
    mb.showinfo('Выберите файл', 'Файл с базой НЕ найден!!')
    file_xlsm = fd.askopenfilename(title='Выберите файл с базой', initialdir=r'C:\Users\User\Емиас\Корректировка ЕГИСЗ')


root = tk.Tk()
root.title("Загрузка прививок в ЭМК")
root.geometry('400x200')
root["bg"] = "#fff"

button1 = tk.Button(text="Обработать загруженный файл",
                    command=read_excel, background="#fff", foreground="#3b3e41",
                    padx="30", pady="15", font="15")



button2 = tk.Button(text="Разделить базу на МО",
                    command=sep_base_mo, background="#fff", foreground="#3b3e41",
                    padx="30", pady="15", font="15")

progress_label = tk.Label(root, text='', font="16")

button1.pack(padx="30", pady="15")
progress_label.pack()
button2.pack(padx="30", pady="15")

root.mainloop()
